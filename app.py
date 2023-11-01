import json
import os
from werkzeug.utils import secure_filename
from flask import Flask, redirect, render_template, request, send_file, session, url_for
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.utils import get_column_letter

env_path = os.path.join(os.path.dirname(__file__), '.env')
load_dotenv(env_path)


def get_meta_q(categoria):
    with open('data/json/metas.json') as file_metas:
        data_metas = json.load(file_metas)
    metas_q = data_metas['Metas']
    for meta in metas_q:
        if meta['categoria'] == categoria:
            return int(meta['cantidad'])
    return ""

def get_meta_monto(categoria):
    with open('data/json/metas.json') as file_metas:
        data_metas = json.load(file_metas)
    metas_q = data_metas['Metas']
    for meta in metas_q:
        if meta['categoria'] == categoria:
            return int(meta['monto'])
    return ""

def descripcion_licencias(legajo):
    with open('data/json/licencias.json') as file_metas:
        data_licencia = json.load(file_metas)
    licencias = data_licencia['Licenses']
    descriptions = []
    for licencia in licencias:
        if licencia['employeeNumber'] == legajo:
            description = licencia['descriptions']
            if description:
                descriptions.append(description)
    if descriptions:
        return ' - '.join(descriptions)
    else:
        return ""
    
def cantidad_licencias(legajo):
    with open('data/json/licencias.json') as file_metas:
        data_licencia = json.load(file_metas)
    licencias = data_licencia['Licenses']
    days = []
    for licencia in licencias:
        if licencia['employeeNumber'] == legajo:
            filter_days = licencia.get('filterDays', 0)
            if filter_days:
                days.append(str(filter_days))
    if days:
        return ' - '.join(days)
    else:
        return ""
  
def read_licencias_especiales():
    with open('data/json/licenciasEspeciales.json') as file_licencias:
        data_licencias = json.load(file_licencias)
    result = []
    for obj in data_licencias['Licencias']:
        result.append(obj)
        legajo = obj['employeeNumber']
    return result  
    
def has_special_licence(legajo):
    licencias_data = read_licencias_especiales()
    for licencia in licencias_data:
        if licencia['employeeNumber'] == legajo:
            return licencia['license']
    return ""

def has_special_licences_days(legajo):
    licencias_data = read_licencias_especiales()
    for licencia in licencias_data:
        if licencia['employeeNumber'] == legajo:
            return int(licencia['licenseDays'])
    return ""    

def read_tutores():
    with open('data/json/tutores.json') as file_tutores:
        data_tutores = json.load(file_tutores)
    result = []
    for obj in data_tutores['Tutores']:
        result.append(obj)
        legajo = obj['legajo']
    return result

def is_tutor(legajo):
    tutores_data = read_tutores()
    tutores_legajos = [tutor['legajo'] for tutor in tutores_data]
    if legajo in tutores_legajos:
        return "SI"
    else:
        return "No"

def read_progresiones():
    with open('data/json/progresionOfUsers.json') as file_progression:
        data_progression = json.load(file_progression)
    result = []
    for obj in data_progression['Progresion_Users']:
        result.append(obj)
        legajo = obj['legajo']
    return result
      
def has_progresiones(legajo):
    progresiones_data = read_progresiones() 
    legajos_con_progresiones = [progresion['legajo'] for progresion in progresiones_data]
    if legajo in legajos_con_progresiones:
        return "SI"
    else:
        return "No"
    
def bimestre_actual():
    mes_actual = datetime.now().month
    bimestre = (mes_actual - 1) // 2 + 1
    mes_inicial = (bimestre - 1) * 2 + 1
    mes_final = bimestre * 2 + 1
    # return list(range(mes_inicial, mes_final))
    return [9,10]

def dias_del_mes_bimestre():
    dias_por_mes = [0, 31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]    
    meses_bimestre = bimestre_actual()
    dias_bimestre = [dias_por_mes[mes] for mes in meses_bimestre]    
    # return dias_bimestre
    return [30,31]

def licencia_mes_uno(legajo, descripcion_lic):
    with open('data/json/licencias.json', 'r') as json_file:
        data = json.load(json_file)
    licencias = data.get('Licenses', []) 
    bim = bimestre_actual()
    dias_bim = dias_del_mes_bimestre()
    dias_licencia = 0
    for licencia in licencias:
        if licencia['employeeNumber'] == legajo and licencia['descriptions'] == descripcion_lic:
            fecha_desde = datetime.strptime(licencia['startDay'], "%d-%m-%Y")
            fecha_hasta = datetime.strptime(licencia['endDay'], "%d-%m-%Y")

            if fecha_desde.month == bim[0] and fecha_hasta.month == bim[0]:
                dias_licencia += (fecha_hasta.day - fecha_desde.day) + 1
            if fecha_desde.month == bim[0] and fecha_hasta.month != bim[0]:
                dias_licencia += (dias_bim[0] - fecha_desde.day) + 1
            if fecha_desde.month != bim[0] and fecha_hasta.month == bim[0]:
                dias_licencia += ((dias_bim[0] + fecha_hasta.day) - dias_bim[0])

    return dias_licencia

def licencia_mes_dos(legajo, descripcion_lic):
    with open('data/json/licencias.json', 'r') as json_file:
        data = json.load(json_file)
    licencias = data.get('Licenses', [])
    bim = bimestre_actual()
    dias_bim = dias_del_mes_bimestre()
    dias_licencia = 0
    for licencia in licencias:
        if licencia['employeeNumber'] == int(legajo) and licencia['descriptions'] == str(descripcion_lic):
            fecha_desde = datetime.strptime(licencia['startDay'], "%d-%m-%Y")
            fecha_hasta = datetime.strptime(licencia['endDay'], "%d-%m-%Y")
            if fecha_desde.month == bim[1] and fecha_hasta.month == bim[1]:
                dias_licencia += (fecha_hasta.day - fecha_desde.day) + 1
            if fecha_desde.month == bim[1] and fecha_hasta.month != bim[1]:
                dias_licencia += (dias_bim[1] - fecha_desde.day) + 1
            if fecha_desde.month != bim[1] and fecha_hasta.month == bim[1]:
                dias_licencia += ((dias_bim[1] + fecha_hasta.day ) - dias_bim[1])
    return dias_licencia

def ajuste_licencias_qm1(legajo):
    with open('data/json/licencias.json', 'r') as json_file:
        data = json.load(json_file)
    licencias = data.get('Licenses', [])
    ajuste_total = 0
    user = next((u for u in licencias if u['employeeNumber'] == legajo), None)
    if user:
        Descripcion_Licencias = user['descriptions']
        if Descripcion_Licencias == "3553-Vacaciones":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "3553-Vacaciones")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0
        if Descripcion_Licencias == "4110-Licencia por Enfermedad":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4110-Licencia por Enfermedad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4070-Licencia por Matrimonio":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4070-Licencia por Matrimonio")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "2222-Licencias por Violencia de genero":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "2222-Licencias por Violencia de genero")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "3333-Licencias por Perdida Gestacional":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "3333-Licencias por Perdida Gestacional")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4104-Licencia por Accidente":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4104-Licencia por Accidente")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                   
        if Descripcion_Licencias == "11-Licencia Fertilidad":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "11-Licencia Fertilidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4132-Licencia por Maternidad":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4132-Licencia por Maternidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4118-Beneficio Paternidad":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4118-Beneficio Paternidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4133-Período de Excedencia":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4133-Período de Excedencia")
            if Cant_Dias_Licencia > 0 and Cant_Dias_Licencia <= 1:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 1:
                ajuste_total += 1.0
            else:
                ajuste_total += 0
    return ajuste_total

def ajuste_licencias_qm2(legajo):
    with open('data/json/licencias.json', 'r') as json_file:
        data = json.load(json_file)
    licencias = data.get('Licenses', [])
    ajuste_total = 0
    user = next((u for u in licencias if u['employeeNumber'] == legajo), None)
    if user:
        Descripcion_Licencias = user['descriptions']
        if Descripcion_Licencias == "3553-Vacaciones":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "3553-Vacaciones")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0
        if Descripcion_Licencias == "4110-Licencia por Enfermedad":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4110-Licencia por Enfermedad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4070-Licencia por Matrimonio":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4070-Licencia por Matrimonio")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "2222-Licencias por Violencia de genero":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "2222-Licencias por Violencia de genero")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "3333-Licencias por Perdida Gestacional":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "3333-Licencias por Perdida Gestacional")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4104-Licencia por Accidente":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4104-Licencia por Accidente")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                   
        if Descripcion_Licencias == "11-Licencia Fertilidad":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "11-Licencia Fertilidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4132-Licencia por Maternidad":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4132-Licencia por Maternidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4118-Beneficio Paternidad":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4118-Beneficio Paternidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4133-Período de Excedencia":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4133-Período de Excedencia")
            if Cant_Dias_Licencia > 0 and Cant_Dias_Licencia <= 1:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 1:
                ajuste_total += 1.0
            else:
                ajuste_total += 0
    return ajuste_total

def ajuste_licencias_monto_uno(legajo):
    with open('data/json/licencias.json', 'r') as json_file:
        data = json.load(json_file)
    licencias = data.get('Licenses', [])
    ajuste_total = 0
    user = next((u for u in licencias if u['employeeNumber'] == legajo), None)
    if user:
        Descripcion_Licencias = user['descriptions']
        if Descripcion_Licencias == "3553-Vacaciones":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "3553-Vacaciones")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0
        if Descripcion_Licencias == "4110-Licencia por Enfermedad":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4110-Licencia por Enfermedad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4070-Licencia por Matrimonio":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4070-Licencia por Matrimonio")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "2222-Licencias por Violencia de genero":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "2222-Licencias por Violencia de genero")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "3333-Licencias por Perdida Gestacional":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "3333-Licencias por Perdida Gestacional")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4104-Licencia por Accidente":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4104-Licencia por Accidente")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                   
        if Descripcion_Licencias == "11-Licencia Fertilidad":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "11-Licencia Fertilidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4132-Licencia por Maternidad":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4132-Licencia por Maternidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4118-Beneficio Paternidad":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4118-Beneficio Paternidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4133-Período de Excedencia":
            Cant_Dias_Licencia = licencia_mes_uno(legajo, "4133-Período de Excedencia")
            if Cant_Dias_Licencia > 0 and Cant_Dias_Licencia <= 1:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 1:
                ajuste_total += 1.0
            else:
                ajuste_total += 0
    return ajuste_total

def ajuste_licencias_monto_dos(legajo):
    with open('data/json/licencias.json', 'r') as json_file:
        data = json.load(json_file)
    licencias = data.get('Licenses', [])
    ajuste_total = 0
    user = next((u for u in licencias if u['employeeNumber'] == legajo), None)
    if user:
        Descripcion_Licencias = user['descriptions']
        if Descripcion_Licencias == "3553-Vacaciones":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "3553-Vacaciones")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0
        if Descripcion_Licencias == "4110-Licencia por Enfermedad":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4110-Licencia por Enfermedad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4070-Licencia por Matrimonio":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4070-Licencia por Matrimonio")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "2222-Licencias por Violencia de genero":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "2222-Licencias por Violencia de genero")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "3333-Licencias por Perdida Gestacional":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "3333-Licencias por Perdida Gestacional")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4104-Licencia por Accidente":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4104-Licencia por Accidente")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                   
        if Descripcion_Licencias == "11-Licencia Fertilidad":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "11-Licencia Fertilidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4132-Licencia por Maternidad":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4132-Licencia por Maternidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4118-Beneficio Paternidad":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4118-Beneficio Paternidad")
            if Cant_Dias_Licencia >= 7 and Cant_Dias_Licencia <=21:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 21:
                ajuste_total += 1.0
            else:
                ajuste_total += 0                    
        if Descripcion_Licencias == "4133-Período de Excedencia":
            Cant_Dias_Licencia = licencia_mes_dos(legajo, "4133-Período de Excedencia")
            if Cant_Dias_Licencia > 0 and Cant_Dias_Licencia <= 1:
                ajuste_total += 0.01 * Cant_Dias_Licencia
            elif Cant_Dias_Licencia > 1:
                ajuste_total += 1.0
            else:
                ajuste_total += 0
    return ajuste_total

def ajuste_meta_q_mes_uno(categoria, legajo):
    ajuste = 0
    def meta_real_q_qm1(categoria):
        with open('data/json/metas.json', 'r') as file_metas:
            data_metas = json.load(file_metas)
            metas = data_metas['Metas']
        for meta in metas:
            if meta['categoria'] == categoria:
                meta_q = meta['cantidad']                
                return int(meta_q)
        return 0
    
    def tutores_ec_qm1(legajo):
        with open('data/json/tutores.json', 'r') as file_tutores:
            data_tutores = json.load(file_tutores)
            tutores = data_tutores['Tutores']
        for tutor in tutores:
            if tutor['legajo'] == legajo:
                return 0.2                    
        return 0
    
    def progresion_ec_qm1(categoria, legajo):
        with open('data/json/progresiones.json', 'r') as file_progresiones:
            data_progresiones = json.load(file_progresiones)
            progresiones = data_progresiones['Progresiones']        
        with open('data/json/progresionOfUsers.json', 'r') as file_progresionOfUsers:
            data_progresionOfUsers = json.load(file_progresionOfUsers)
            progresionOfUsers = data_progresionOfUsers['Progresion_Users']        
            for user in progresionOfUsers:
                if user['legajo'] == legajo:
                    mes_progresion = user['mes_progresion']
                    for progresion in progresiones:
                        if progresion['categoria'] == categoria:
                            ajuste = progresion[mes_progresion]
                            return ajuste
        return 0  
    
    def ajuste_licencia_especial_qm1(legajo):
        with open('data/json/licenciasEspeciales.json', 'r') as file_licencia:
            data_licencia = json.load(file_licencia)
            licencia = data_licencia['Licencias']
        for user in licencia:
            if user['employeeNumber'] == legajo and 'adjustment' in user:
                return float(user['adjustment'])        
        return 0.0
    
    meta_real = meta_real_q_qm1(categoria)
    ajuste_de_licencias = ajuste_licencias_qm1(legajo)
    ajuste_de_licencia_especial = ajuste_licencia_especial_qm1(legajo)
    ajuste_tutores = tutores_ec_qm1(legajo)
    ajuste_progresion = progresion_ec_qm1(categoria, legajo)
    ajuste_total = ajuste_tutores + ajuste_progresion + ajuste_de_licencias + ajuste_de_licencia_especial
    valor_maximo = max(ajuste_de_licencias, ajuste_de_licencia_especial, ajuste_tutores)
    if valor_maximo == ajuste_de_licencias:
        segundo_mayor = max(ajuste_de_licencia_especial, ajuste_tutores)
        menor = min(ajuste_de_licencia_especial, ajuste_tutores)
    elif valor_maximo == ajuste_de_licencia_especial:
        segundo_mayor = max(ajuste_de_licencias, ajuste_tutores)
        menor = min(ajuste_de_licencias, ajuste_tutores)
    else:
        segundo_mayor = max(ajuste_de_licencias, ajuste_de_licencia_especial)
        menor = min(ajuste_de_licencias, ajuste_de_licencia_especial)
    ajuste_uno = meta_real - (meta_real*ajuste_progresion)
    ajuste_dos = ajuste_uno - (ajuste_uno * valor_maximo)
    ajuste_tres = ajuste_dos - (ajuste_dos*segundo_mayor)
    ajuste += ajuste_tres - (ajuste_tres * menor)
    if ajuste_total >= 1:
        ajuste = 0  
    return float(round(ajuste, 2))

def ajuste_meta_q_mes_dos(categoria, legajo):
    ajuste = 0
    def meta_real_q_qm1(categoria):
        with open('data/json/metas.json', 'r') as file_metas:
            data_metas = json.load(file_metas)
            metas = data_metas['Metas']
        for meta in metas:
            if meta['categoria'] == categoria:
                meta_q = meta['cantidad']                
                return int(meta_q)
        return 0
    
    def tutores_ec_qm1(legajo):
        with open('data/json/tutores.json', 'r') as file_tutores:
            data_tutores = json.load(file_tutores)
            tutores = data_tutores['Tutores']
        for tutor in tutores:
            if tutor['legajo'] == legajo:
                return 0.2                    
        return 0
    
    def progresion_ec_qm1(categoria, legajo):
        with open('data/json/progresiones.json', 'r') as file_progresiones:
            data_progresiones = json.load(file_progresiones)
            progresiones = data_progresiones['Progresiones']        
        with open('data/json/progresionOfUsers.json', 'r') as file_progresionOfUsers:
            data_progresionOfUsers = json.load(file_progresionOfUsers)
            progresionOfUsers = data_progresionOfUsers['Progresion_Users']        
            for user in progresionOfUsers:
                if user['legajo'] == legajo:
                    mes_progresion = user['mes_progresion']
                    for progresion in progresiones:
                        if progresion['categoria'] == categoria:
                            ajuste = progresion[mes_progresion]
                            return ajuste
        return 0
    
    def ajuste_licencia_especial_qm1(legajo):
        with open('data/json/licenciasEspeciales.json', 'r') as file_licencia:
            data_licencia = json.load(file_licencia)
            licencia = data_licencia['Licencias']
        for user in licencia:
            if user['employeeNumber'] == legajo and 'adjustment' in user:
                return float(user['adjustment'])        
        return 0.0
    
    meta_real = meta_real_q_qm1(categoria)
    ajuste_de_licencias = ajuste_licencias_qm2(legajo)
    ajuste_de_licencia_especial = ajuste_licencia_especial_qm1(legajo)
    ajuste_tutores = tutores_ec_qm1(legajo)
    ajuste_progresion = progresion_ec_qm1(categoria, legajo)
    ajuste_total = ajuste_tutores + ajuste_progresion + ajuste_de_licencias + ajuste_de_licencia_especial
    valor_maximo = max(ajuste_de_licencias, ajuste_de_licencia_especial, ajuste_tutores)
    if valor_maximo == ajuste_de_licencias:
        segundo_mayor = max(ajuste_de_licencia_especial, ajuste_tutores)
        menor = min(ajuste_de_licencia_especial, ajuste_tutores)
    elif valor_maximo == ajuste_de_licencia_especial:
        segundo_mayor = max(ajuste_de_licencias, ajuste_tutores)
        menor = min(ajuste_de_licencias, ajuste_tutores)
    else:
        segundo_mayor = max(ajuste_de_licencias, ajuste_de_licencia_especial)
        menor = min(ajuste_de_licencias, ajuste_de_licencia_especial)
    ajuste_uno = meta_real - (meta_real*ajuste_progresion)
    ajuste_dos = ajuste_uno - (ajuste_uno * valor_maximo)
    ajuste_tres = ajuste_dos - (ajuste_dos*segundo_mayor)
    ajuste += ajuste_tres - (ajuste_tres * menor)
    if ajuste_total >= 1:
        ajuste = 0
    return float(round(ajuste, 2))

def ajuste_meta_monto_m1(categoria, legajo):
    ajuste = 0
    def meta_real_monto(categoria):
        with open('data/json/metas.json', 'r') as file_metas:
            data_metas = json.load(file_metas)
            metas = data_metas['Metas']
        for meta in metas:
            if meta['categoria'] == categoria:
                meta_monto = meta['monto']
                return meta_monto
        return 0

    def tutores_ec(legajo):
        with open('data/json/tutores.json', 'r') as file_tutores:
            data_tutores = json.load(file_tutores)
            tutores = data_tutores['Tutores']
        for tutor in tutores:
            if tutor['legajo'] == legajo:
                    return 0.2                
        return 0

    def progresion_ec(categoria, legajo):  
        with open('data/json/progresiones.json', 'r') as file_progresiones:
            data_progresiones = json.load(file_progresiones)
            progresiones = data_progresiones['Progresiones']        
        with open('data/json/progresionOfUsers.json', 'r') as file_progresionOfUsers:
            data_progresionOfUsers = json.load(file_progresionOfUsers)
            progresionOfUsers = data_progresionOfUsers['Progresion_Users']        
            for user in progresionOfUsers:
                if user['legajo'] == legajo:
                    mes_progresion = user['mes_progresion']
                    for progresion in progresiones:
                        if progresion['categoria'] == categoria:
                            ajuste = progresion[mes_progresion]
                            return ajuste
        return 0
    
    def ajuste_licencia_especial(legajo):
        with open('data/json/licenciasEspeciales.json', 'r') as file_licencia:
            data_licencia = json.load(file_licencia)
            licencia = data_licencia['Licencias']
        for user in licencia:
            if user['employeeNumber'] == legajo and 'adjustment' in user:
                return float(user['adjustment'])        
        return 0.0

    meta_real = meta_real_monto(categoria)
    ajuste_de_licencias = ajuste_licencias_monto_uno(legajo)
    ajuste_de_licencia_especial = ajuste_licencia_especial(legajo)
    ajuste_tutores = tutores_ec(legajo)
    ajuste_progresion = progresion_ec(categoria, legajo)
    ajuste_total = ajuste_tutores + ajuste_progresion + ajuste_de_licencias + ajuste_de_licencia_especial
    valor_maximo = max(ajuste_de_licencias, ajuste_de_licencia_especial, ajuste_tutores)
    if valor_maximo == ajuste_de_licencias:
        segundo_mayor = max(ajuste_de_licencia_especial, ajuste_tutores)
        menor = min(ajuste_de_licencia_especial, ajuste_tutores)
    elif valor_maximo == ajuste_de_licencia_especial:
        segundo_mayor = max(ajuste_de_licencias, ajuste_tutores)
        menor = min(ajuste_de_licencias, ajuste_tutores)
    else:
        segundo_mayor = max(ajuste_de_licencias, ajuste_de_licencia_especial)
        menor = min(ajuste_de_licencias, ajuste_de_licencia_especial)
    ajuste_uno = meta_real - (meta_real*ajuste_progresion)
    ajuste_dos = ajuste_uno - (ajuste_uno * valor_maximo)
    ajuste_tres = ajuste_dos - (ajuste_dos*segundo_mayor)
    ajuste += ajuste_tres - (ajuste_tres * menor)
    if ajuste_total >= 1:
        ajuste = 0  
    return int(ajuste)

def ajuste_meta_monto_m2(categoria, legajo):
    ajuste = 0
    def meta_real_monto(categoria):
        with open('data/json/metas.json', 'r') as file_metas:
            data_metas = json.load(file_metas)
            metas = data_metas['Metas']
        for meta in metas:
            if meta['categoria'] == categoria:
                meta_monto = meta['monto']
                return meta_monto
        return 0

    def tutores_ec(legajo):
        with open('data/json/tutores.json', 'r') as file_tutores:
            data_tutores = json.load(file_tutores)
            tutores = data_tutores['Tutores']
        for tutor in tutores:
            if tutor['legajo'] == legajo:
                    return 0.2                
        return 0

    def progresion_ec(categoria, legajo):  
        with open('data/json/progresiones.json', 'r') as file_progresiones:
            data_progresiones = json.load(file_progresiones)
            progresiones = data_progresiones['Progresiones']        
        with open('data/json/progresionOfUsers.json', 'r') as file_progresionOfUsers:
            data_progresionOfUsers = json.load(file_progresionOfUsers)
            progresionOfUsers = data_progresionOfUsers['Progresion_Users']        
            for user in progresionOfUsers:
                if user['legajo'] == legajo:
                    mes_progresion = user['mes_progresion']
                    for progresion in progresiones:
                        if progresion['categoria'] == categoria:
                            ajuste = progresion[mes_progresion]
                            return ajuste
        return 0
    
    def ajuste_licencia_especial(legajo):
        with open('data/json/licenciasEspeciales.json', 'r') as file_licencia:
            data_licencia = json.load(file_licencia)
            licencia = data_licencia['Licencias']
        for user in licencia:
            if user['employeeNumber'] == legajo and 'adjustment' in user:
                return float(user['adjustment'])        
        return 0.0

    meta_real = meta_real_monto(categoria)
    ajuste_de_licencias = ajuste_licencias_monto_dos(legajo)
    ajuste_de_licencia_especial = ajuste_licencia_especial(legajo)
    ajuste_tutores = tutores_ec(legajo)
    ajuste_progresion = progresion_ec(categoria, legajo)
    ajuste_total = ajuste_tutores + ajuste_progresion + ajuste_de_licencias + ajuste_de_licencia_especial
    valor_maximo = max(ajuste_de_licencias, ajuste_de_licencia_especial, ajuste_tutores)
    if valor_maximo == ajuste_de_licencias:
        segundo_mayor = max(ajuste_de_licencia_especial, ajuste_tutores)
        menor = min(ajuste_de_licencia_especial, ajuste_tutores)
    elif valor_maximo == ajuste_de_licencia_especial:
        segundo_mayor = max(ajuste_de_licencias, ajuste_tutores)
        menor = min(ajuste_de_licencias, ajuste_tutores)
    else:
        segundo_mayor = max(ajuste_de_licencias, ajuste_de_licencia_especial)
        menor = min(ajuste_de_licencias, ajuste_de_licencia_especial)
    ajuste_uno = meta_real - (meta_real*ajuste_progresion)
    ajuste_dos = ajuste_uno - (ajuste_uno * valor_maximo)
    ajuste_tres = ajuste_dos - (ajuste_dos*segundo_mayor)
    ajuste += ajuste_tres - (ajuste_tres * menor)
    if ajuste_total >= 1:
        ajuste = 0  
    return int(ajuste)

def ajuste_total_q(categoria, legajo):
    ajuste_q_mes_1 = ajuste_meta_q_mes_uno(categoria, legajo)
    ajuste_q_mes_2 = ajuste_meta_q_mes_dos(categoria, legajo)
    result = (ajuste_q_mes_1/2) + (ajuste_q_mes_2/2)    
    return float(round(result,2))

def ajuste_total_monto(categoria, legajo):
    ajuste_monto_mes_1 = ajuste_meta_monto_m1(categoria, legajo)
    ajuste_monto_mes_2 = ajuste_meta_monto_m2(categoria, legajo)
    result = (ajuste_monto_mes_1/2) + (ajuste_monto_mes_2/2)    
    return float(result)

def serialize_datetime(obj):
    if isinstance(obj, datetime):
        return obj.strftime('%d-%m-%Y')
    raise TypeError("Tipo no serializable")

def convert_excel_licencia__to_json(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)
    json_data = {"Licenses": data}
    return json_data

def convert_excel_to_json(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        row_data["Observacion"] = ""
        data.append(row_data)
    json_data = {"Nomina": data}
    return json_data

def export_to_excel(nomina_data):
    wb = Workbook()
    ws = wb.active
    bold_font = Font(bold=True)
    light_green = Color(rgb="C6EFCE")
    fill = PatternFill(start_color=light_green, end_color=light_green, fill_type="solid")
    columns = ["Legajo", "Nombre Ejecutivo", "Sucursal", "Categoria", "Meta Q", "Meta $", "Descripción Licencias", "Dias Licencencia", "Licencias Especiales", "Dias Licencias Especiales", "Es Tutor", "Tiene Progresion", "Ajuste Q, 1° Mes", "Ajuste Q, 2° Mes", "Ajuste Monto, 1° Mes", "Ajuste Monto, 2° Mes", "Ajuste Total Q", "Ajuste total Monto", "Observaciones"]
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font
        cell.fill = fill
    for user in nomina_data:
        row_data = [
            user.get("employeeNumber", ""),
            user.get("fullName", ""),
            user.get("branch", ""),
            user.get("category", ""),
            get_meta_q(user.get("category", "")),
            get_meta_monto(user.get("category", "")),
            descripcion_licencias(user.get("employeeNumber", "")),
            cantidad_licencias(user.get("employeeNumber", "")),
            has_special_licence(user.get("employeeNumber", "")),
            has_special_licences_days(user.get("employeeNumber", "")),
            is_tutor(user.get("employeeNumber", "")),
            has_progresiones(user.get("employeeNumber", "")),
            ajuste_meta_q_mes_uno(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_q_mes_dos(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_monto_m1(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_monto_m2(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_total_q(user.get("category", ""), user.get("employeeNumber", "")),
            ajuste_total_monto(user.get("category", ""), user.get("employeeNumber", "")),
            user.get("Observacion", "")
        ]
        ws.append(row_data)
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(len(str(cell.value)) for cell in ws[col_letter])
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width    
    # current_dir = os.path.dirname(os.path.abspath(__file__))
    # excel_file_path = os.path.join(current_dir, 'data', 'export', 'Validacion_Metas.xlsx')
    excel_file_path = os.path.join('data/export/Validacion_Metas.xlsx')
    excel_file = excel_file_path
    wb.save(excel_file)
    return excel_file

def export_to_excel_azure(nomina_data):
    wb = Workbook()
    ws = wb.active
    bold_font = Font(bold=True)
    light_green = Color(rgb="C6EFCE")
    fill = PatternFill(start_color=light_green, end_color=light_green, fill_type="solid")
    columns = ["Legajo_EC", "Meta_Q", "Meta_Monto", "Meta_Q_MES_1", "Meta_Monto_MES_1", "Meta_Q_MES_2", "Meta_Monto_MES_2", "Mes_1", "Mes_2", "Anio"]
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font
        cell.fill = fill
    for user in nomina_data:
        row_data = [
            user.get("employeeNumber", ""),
            ajuste_total_q(user.get("category", ""), user.get("employeeNumber", "")),
            ajuste_total_monto(user.get("category", ""), user.get("employeeNumber", "")),
            ajuste_meta_q_mes_uno(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_monto_m1(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_q_mes_dos(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_monto_m2(user.get("category", ""), user.get("employeeNumber", ""))/2,
            bimestre_actual()[0],
            bimestre_actual()[1],
            datetime.now().year
        ]
        ws.append(row_data)
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(len(str(cell.value)) for cell in ws[col_letter])
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width
    fecha = str(datetime.now().strftime("%d-%m-%Y"))  
    excel_file_path = os.path.join(f'data/export/Metas Colocaciones {fecha}.xlsx')
    excel_file = excel_file_path
    wb.save(excel_file)
    return excel_file

def handle_form_submission(jefe_zonal, nomina):
    with open(nomina, 'r') as file_nomina:
        data_metas = json.load(file_nomina)
        nomina = data_metas['Nomina']    
    filtered_nomina = [entry for entry in nomina if entry['jefe_a_cargo'] == jefe_zonal]
    return filtered_nomina


app = Flask(__name__, static_folder='public')
app.secret_key = 'tu_clave_secreta'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# TODO----------LOGIN--------------------------------------------------------------------------------------------------
@app.route('/', methods=['GET', 'POST'])
def login():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'users.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        usuarios = data['Users']    
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        for user in usuarios:
            if user['mail'] == email:
                if user['password'] == password:
                    session['usuario'] = user
                    return redirect(url_for('index'))
                else:
                    session['error_message'] = 'La contraseña es incorrecta'
                    return redirect(url_for('login'))        
        return render_template('not_authorized.html') 
    error_message = session.pop('error_message', None)   
    return render_template('login.html', 
                           error_message=error_message)

@app.route('/index', methods=['GET', 'POST'])
def index():
    usuario = session.get('usuario')
    if usuario:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        json_path = os.path.join(current_dir, 'data', 'json', 'users.json')
        
        with open(json_path, 'r') as json_file:
            data = json.load(json_file)
        usuarios = data['Users']
        perfil = next((usr['perfil'] for usr in usuarios if usr['legajo'] == usuario['legajo']), None)
        nombre = usuario['nombre']
        return render_template('index.html', 
                               nombre=nombre, 
                               perfil=perfil)
      
@app.route('/logout', methods=['POST'])
def logout():
    session.pop('usuario', None)
    return redirect(url_for('login'))

# TODO----------NOMINA-------------------------------------------------------------------------------------------------
@app.route('/nomina', methods=['GET', 'POST'])
def nomina():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'nomina.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        nomina = data["Nomina"]
    
    if request.method == 'POST':
        jefe_zonal = request.form['user']
        filtered_nomina = handle_form_submission(jefe_zonal)
        return render_template('nomina.html', 
                               nomina=filtered_nomina, 
                               is_tutor=is_tutor, 
                               has_progresiones=has_progresiones, 
                               get_meta_q=get_meta_q, 
                               get_meta_monto=get_meta_monto, 
                               has_special_licence=has_special_licence, 
                               has_special_licences_days=has_special_licences_days, 
                               handle_form_submission=handle_form_submission)
    else:
        return render_template('nomina.html', 
                               nomina=nomina, 
                               is_tutor=is_tutor, 
                               has_progresiones=has_progresiones, 
                               get_meta_q=get_meta_q, 
                               get_meta_monto=get_meta_monto, 
                               has_special_licence=has_special_licence, 
                               has_special_licences_days=has_special_licences_days, 
                               ajuste_meta_q_mes_uno=ajuste_meta_q_mes_uno, 
                               ajuste_meta_q_mes_dos=ajuste_meta_q_mes_dos, 
                               ajuste_meta_monto_m1=ajuste_meta_monto_m1, 
                               ajuste_meta_monto_m2=ajuste_meta_monto_m2, 
                               cantidad_licencias=cantidad_licencias, 
                               descripcion_licencias=descripcion_licencias, 
                               ajuste_total_q=ajuste_total_q, 
                               ajuste_total_monto=ajuste_total_monto)

@app.route('/edit_observacion/<legajo>', methods=['GET', 'POST'])
def edit_observacion(legajo):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'nomina.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    nomina = data['Nomina']
    usuario = next((user for user in nomina if user['employeeNumber'] == int(legajo)), None)
    if request.method == 'POST':
        nueva_observacion = request.form.get('Observacion')
        usuario['Observacion'] = nueva_observacion
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
        return redirect(url_for('nomina'))
    
    return render_template('edit/editObservacion.html', 
                           usuario=usuario)

@app.route('/delete_observacion', methods=['POST'])
def delete_observacion():
    legajo = request.form['employeeNumber']
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'nomina.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)    
    nomina = data['Nomina']
    usuario = next((user for user in nomina if user['employeeNumber'] == int(legajo)), None)
    if usuario:
        usuario['Observacion'] = ""
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)            
    return redirect('/nomina')

# TODO----------Licencias-------------------------------------------------------------------------------------------------
@app.route('/licencias', methods=['GET', 'POST'])
def licencias():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'licencias.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    licencias = []
    for licencia in data['Licenses']:
        start_day = datetime.strptime(licencia['startDay'], "%d-%m-%Y")
        end_day = datetime.strptime(licencia['endDay'], "%d-%m-%Y")
        licencias.append({
            "employeeNumber": licencia['employeeNumber'],
            "fullName": licencia['fullName'],
            "descriptions": licencia['descriptions'],
            "startDay": start_day.strftime("%d-%m-%Y"),
            "endDay": end_day.strftime("%d-%m-%Y"),
            "filterDays": licencia['filterDays'],
            "licenseDays": licencia['licenseDays']
        })        
    return render_template('licencias.html', 
                           licencias=licencias)

# TODO-----------METAS------------------------------------------------------------------------------------------------
@app.route('/metas', methods=['GET', 'POST'])
def metas():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'metas.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        metas = data["Metas"]
    return render_template('metas.html', 
                           metas=metas)

@app.route('/editar-meta/<categoria>', methods=['GET', 'POST'])
def editar_meta(categoria):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'metas.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    metas = data['Metas']
    datos_categoria = next((meta for meta in metas if meta['categoria'] == categoria), None)    
    if request.method == 'POST':
        nueva_cantidad = int(request.form.get('cantidad'))
        nuevo_monto = float(request.form.get('monto'))
        nuevo_monto_prom = float(request.form.get('monto_prom'))
        nueva_fecha = request.form.get('fecha')
        datos_categoria['cantidad'] = nueva_cantidad
        datos_categoria['monto'] = nuevo_monto
        datos_categoria['monto_prom'] = nuevo_monto_prom
        datos_categoria['fecha'] = nueva_fecha        
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)        
        return redirect(url_for('metas'))
    return render_template('edit/editMetas.html', 
                           datos_categoria=datos_categoria)

@app.route('/eliminar-meta/<categoria>', methods=['POST'])
def eliminar_meta(categoria):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'metas.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    metas = data['Metas']
    meta_a_eliminar = None
    for meta in metas:
        if meta['categoria'] == categoria:
            meta_a_eliminar = meta
            break
    if meta_a_eliminar:
        metas.remove(meta_a_eliminar)
    with open(json_path, 'w') as json_file:
        json.dump(data, json_file, indent=2)
    return redirect(url_for('metas'))

@app.route('/agregar-meta', methods=['GET', 'POST'])
def agregar_meta():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'metas.json')    
    if request.method == 'POST':
        categoria = request.form['categoria']
        cantidad = int(request.form['cantidad'])
        monto = float(request.form['monto'])
        monto_prom = float(request.form['monto_prom'])
        fecha = request.form['fecha']        
        with open(json_path, 'r') as json_file:
            data = json.load(json_file)
            metas = data['Metas']
        if any(meta['categoria'] == categoria for meta in metas):
            error_message = 'La categoría ya existe. No se permite agregar una meta con la misma categoría.'
            return render_template('add/addMetas.html', error_message=error_message)        
        nueva_meta = {
            'categoria': categoria,
            'cantidad': cantidad,
            'monto': monto,
            'monto_prom': monto_prom,
            'fecha': fecha
        }        
        metas.append(nueva_meta)        
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)        
        return redirect(url_for('metas'))    
    return render_template('add/addMetas.html')

# TODO-----------PROGRESIONES-----------------------------------------------------------------------------------------
@app.route('/progresiones', methods=['GET', 'POST'])
def progresiones():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'progresiones.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        progresiones = data["Progresiones"]
    return render_template('progresiones.html', 
                           progresiones=progresiones)

@app.route('/editar-progresion/<categoria>', methods=['GET', 'POST'])
def editar_progresion(categoria):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'progresiones.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    progresiones = data['Progresiones']
    datos_progresion = next((progresion for progresion in progresiones if progresion['categoria'] == categoria), None)    
    if request.method == 'POST':
        datos_progresion['mes_1'] = float(request.form.get('mes_1'))
        datos_progresion['mes_2'] = float(request.form.get('mes_2'))
        datos_progresion['mes_3'] = float(request.form.get('mes_3'))
        datos_progresion['mes_4'] = float(request.form.get('mes_4'))
        datos_progresion['mes_5'] = float(request.form.get('mes_5'))
        datos_progresion['mes_6'] = float(request.form.get('mes_6'))              
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)        
        return redirect(url_for('progresiones'))
    return render_template('edit/editProgresiones.html', 
                           datos_progresion=datos_progresion)

@app.route('/eliminar-progresion/<categoria>', methods=['POST'])
def eliminar_progresion(categoria):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'progresiones.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    progresiones = data['Progresiones']
    progresion_a_eliminar = None
    for progresion in progresiones:
        if progresion['categoria'] == categoria:
            progresion_a_eliminar = progresion
            break
    if progresion_a_eliminar:
        progresiones.remove(progresion_a_eliminar)
    with open(json_path, 'w') as json_file:
        json.dump(data, json_file, indent=2)
    return redirect(url_for('progresiones'))

@app.route('/agregar-progresion', methods=['GET', 'POST'])
def agregar_progresion():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'progresiones.json')    
    if request.method == 'POST':
        categoria = request.form['categoria']
        mes_1 = float(request.form['mes_1'])
        mes_2 = float(request.form['mes_2'])
        mes_3 = float(request.form['mes_3'])
        mes_4 = float(request.form['mes_4'])
        mes_5 = float(request.form['mes_5'])
        mes_6 = float(request.form['mes_6'])        
        with open(json_path, 'r') as json_file:
            data = json.load(json_file)
            progresiones = data['Progresiones']   
        if any(progresion['categoria'] == categoria for progresion in progresiones):
            error_message = 'La categoría ya existe. No se permite agregar otra progresion con la misma categoría.'
            return render_template('add/addMetas.html', 
                                   error_message=error_message)       
        nueva_progresion = {
            'categoria': categoria,
            'mes_1': mes_1,
            'mes_2': mes_2,
            'mes_3': mes_3,
            'mes_4': mes_4,
            'mes_5': mes_5,
            'mes_6': mes_6
        }        
        progresiones.append(nueva_progresion)        
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)        
        return redirect(url_for('progresiones'))    
    return render_template('add/addProgresiones.html')

# TODO------------USERS------------------------------------------------------------------------------------------
@app.route('/users', methods=['GET', 'POST'])
def users():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'users.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        users = data["Users"]        
    selected_user = request.form.get('user')
    return render_template('users.html', 
                           users=users, 
                           selected_user=selected_user)

@app.route('/eliminar-user/<legajo>', methods=['POST'])
def eliminar_user(legajo):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'users.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    users = data['Users']    
    user_a_eliminar = None
    for user in users:
        if user['legajo'] == int(legajo):
            user_a_eliminar = user
            break        
    if user_a_eliminar:
        users.remove(user_a_eliminar)    
    with open(json_path, 'w') as json_file:
        json.dump(data, json_file, indent=2)        
    return redirect(url_for('users'))    
               
@app.route('/editar-usuario/<legajo>', methods=['GET', 'POST'])
def editar_usuario(legajo):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'users.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    usuarios = data['Users']
    datos_usuario = next((usuario for usuario in usuarios if usuario['legajo'] == int(legajo)), None)    
    if request.method == 'POST':
        nuevo_nombre = request.form.get('nombre')
        nuevo_mail = request.form.get('mail')
        nuevo_passsw = request.form.get('password')
        nuevo_perfil = request.form.get('perfil')
        datos_usuario['nombre'] = nuevo_nombre
        datos_usuario['mail'] = nuevo_mail
        datos_usuario['password'] = nuevo_passsw
        datos_usuario['perfil'] = nuevo_perfil        
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)        
        return redirect(url_for('users'))
    return render_template('edit/editUsers.html',
                           datos_usuario=datos_usuario)

@app.route('/agregar-usuario', methods=['GET', 'POST'])
def agregar_usuario():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'users.json')
    if request.method == 'POST':
        legajo = int(request.form['legajo'])
        nombre = request.form['nombre']
        mail = request.form['mail']
        password = request.form['password']
        perfil = request.form['perfil']
        with open(json_path, 'r') as json_file:
            data = json.load(json_file)
            usuarios = data['Users']
        if any(usuario['legajo'] == legajo for usuario in usuarios):
            error_message = 'El legajo ya existe. No se permite agregar un usuario con el mismo legajo.'
            return render_template('add/addUsers.html', error_message=error_message)        
        nuevo_usuario = {
            'legajo': legajo,
            'nombre': nombre,
            'mail': mail,
            'password': password,
            'perfil': perfil
        }
        usuarios.append(nuevo_usuario)
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
        return redirect(url_for('users'))
    return render_template('add/addUsers.html')

# TODO-------------USUARIOS_CON_PROGRESIONES---------------------------------------------------------------------------
@app.route('/usuariosConProgresiones')
def usuarios_con_progresion():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'progresionOfUsers.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        usuarios = data['Progresion_Users']
    return render_template('usuariosConProgresiones.html', 
                           usuarios=usuarios)

@app.route('/editar-usuarioConProgresiones/<legajo>', methods=['GET', 'POST'])
def editar_usuario_con_progresiones(legajo):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'progresionOfUsers.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    usuarios = data["Progresion_Users"]
    datos_usuarioConProgresiones = next((usuario for usuario in usuarios if usuario['legajo'] == int(legajo)), None)
    if request.method == 'POST':
        datos_usuarioConProgresiones['apellido'] = request.form.get('apellido')
        datos_usuarioConProgresiones['nombre'] = request.form.get('nombre')
        datos_usuarioConProgresiones['mail'] = request.form.get('mail')
        datos_usuarioConProgresiones['categoria'] = request.form.get('categoria')
        datos_usuarioConProgresiones['fecha_desde'] = request.form.get('fecha_desde')
        datos_usuarioConProgresiones['fecha_hasta'] = request.form.get('fecha_hasta')
        datos_usuarioConProgresiones['mes_progresion'] = request.form.get('mes_progresion')
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
        return redirect(url_for('usuarios_con_progresion'))
    return render_template('edit/editUsuariosConProgresiones.html', 
                           datos_usuarioConProgresiones=datos_usuarioConProgresiones)

@app.route('/eliminar-usuarioConProgresiones/<legajo>', methods=['POST'])
def eliminar_usuario_con_progresiones(legajo):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'progresionOfUsers.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    usuarios = data['Progresion_Users']
    usuario_a_eliminar = next((usuario for usuario in usuarios if usuario['legajo'] == int(legajo)), None)
    if usuario_a_eliminar:
        usuarios.remove(usuario_a_eliminar)
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
    return redirect(url_for('usuarios_con_progresion'))

@app.route('/agregar-usuarioConProgresion', methods=['GET', 'POST'])
def agregar_usuario_con_progresion():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'progresionOfUsers.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        usuarios = data['Progresion_Users']
    if request.method == 'POST':
        legajo = int(request.form.get('legajo'))
        
        # Validar si el legajo ya existe en el archivo JSON
        if any(usuario['legajo'] == legajo for usuario in usuarios):
            error_message = 'El legajo ya existe. No se permite agregar un usuario con el mismo legajo.'
            return render_template('add/addUsuarioConProgresion.html', 
                                   error_message=error_message)
        
        nuevo_usuario = {
            'legajo': legajo,
            'apellido': request.form.get('apellido'),
            'nombre': request.form.get('nombre'),
            'mail': request.form.get('mail'),
            'categoria': request.form.get('categoria'),
            'fecha_desde': request.form.get('fecha_desde'),
            'fecha_hasta': request.form.get('fecha_hasta'),
            'mes_progresion': request.form.get('mes_progresion')
        }
        usuarios.append(nuevo_usuario)
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
        return redirect(url_for('usuarios_con_progresion'))
    return render_template('add/addUsuarioConProgresion.html', 
                           error_message=None)

# TODO-------------TUTORES------------------------------------------------------------------------------------
@app.route('/tutores')
def tutores():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'tutores.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        usuarios = data['Tutores']
    return render_template('tutores.html', 
                           usuarios=usuarios)

@app.route('/editar-Tutores/<legajo>', methods=['GET', 'POST'])
def editar_tutores(legajo):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'tutores.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        usuarios = data['Tutores']
    datos_tutores = next((usuario for usuario in usuarios if usuario['legajo'] == int(legajo)), None)
    if request.method == 'POST':
        datos_tutores['nombre'] = request.form.get('nombre')
        datos_tutores['apellido'] = request.form.get('apellido')
        datos_tutores['categoria'] = request.form.get('categoria')
        datos_tutores['fecha_inicio_tutoria'] = request.form.get('fecha_inicio_tutoria')
        datos_tutores['fecha_fin_tutoria'] = request.form.get('fecha_fin_tutoria')
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
        return redirect(url_for('tutores'))
    return render_template('edit/editTutores.html', 
                           datos_tutores=datos_tutores)
         
@app.route('/eliminar-tutores/<legajo>', methods=['POST'])
def eliminar_tutores(legajo):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'tutores.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    usuarios = data['Tutores']
    tutor_a_eliminar = next((usuario for usuario in usuarios if usuario['legajo'] == int(legajo)), None)
    if tutor_a_eliminar:
        usuarios.remove(tutor_a_eliminar)
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
    return redirect(url_for('tutores'))

@app.route('/agregar-Tutor', methods=['GET', 'POST'])
def agregar_tutor():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'tutores.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        tutores = data['Tutores']
    if request.method == 'POST':
        legajo = int(request.form.get('legajo'))       
        # Validar si el legajo ya existe en el archivo JSON
        if any(tutor['legajo'] == legajo for tutor in tutores):
            error_message = 'El legajo ya existe. No se permite agregar un tutor con el mismo legajo.'
            return render_template('add/addTutores.html', error_message=error_message)        
        nuevo_tutor = {
            'legajo': legajo,
            'nombre': request.form.get('nombre'),
            'apellido': request.form.get('apellido'),
            'categoria': request.form.get('categoria'),
            'fecha_inicio_tutoria': request.form.get('fecha_inicio_tutoria'),
            'fecha_fin_tutoria': request.form.get('fecha_fin_tutoria')
        }
        tutores.append(nuevo_tutor)
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
        return redirect(url_for('tutores'))
    return render_template('add/addTutores.html',
                           error_message=None)

#TODO--------LICENCIAS ESPECIALES--------------------------------------------------------
@app.route('/licenciasEspeciales')
def licencias_especiales():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'licenciasEspeciales.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        licencias = data['Licencias']
    return render_template('licenciasEspeciales.html', 
                           licencias=licencias)

@app.route('/editar-Licencia/<employeeNumber>', methods=['GET', 'POST'])
def editar_licencia(employeeNumber):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'licenciasEspeciales.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        usuarios = data['Licencias']
    datos_licencia = next((usuario for usuario in usuarios if usuario['employeeNumber'] == int(employeeNumber)), None)
    if request.method == 'POST':
        datos_licencia['fullName'] = request.form.get('fullName')
        datos_licencia['license'] = request.form.get('license')
        datos_licencia['licenseStar'] = request.form.get('licenseStar')
        datos_licencia['licenseEnd'] = request.form.get('licenseEnd')
        datos_licencia['licenseDays'] = request.form.get('licenseDays')
        datos_licencia['adjustment'] = float(request.form.get('adjustment'))
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
        return redirect(url_for('licencias_especiales'))
    return render_template('edit/editLicencias.html', 
                           datos_licencia=datos_licencia)

@app.route('/eliminar-Licencia/<employeeNumber>', methods=['POST'])
def eliminar_licencia(employeeNumber):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'licenciasEspeciales.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
    usuarios = data['Licencias']
    licencia_a_eliminar = next((usuario for usuario in usuarios if usuario['employeeNumber'] == int(employeeNumber)), None)
    if licencia_a_eliminar:
        usuarios.remove(licencia_a_eliminar)
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
    return redirect(url_for('licencias_especiales'))

@app.route('/agregar-Licencia', methods=['GET', 'POST'])
def agregar_licencia():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'licenciasEspeciales.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        licencias = data['Licencias']
    if request.method == 'POST':
        employeeNumber = int(request.form.get('employeeNumber'))
        if any(licen['employeeNumber'] == employeeNumber for licen in licencias):
            error_message = 'El legajo ya existe. No se permite agregar una licencia con el mismo legajo.'
            return render_template('add/addLicencias.html', error_message=error_message)        
        nueva_licencia = {
            'employeeNumber': employeeNumber,
            'fullName': request.form.get('fullName'),
            'license': request.form.get('license'),
            'licenseStar': request.form.get('licenseStar'),
            'licenseEnd': request.form.get('licenseEnd'),
            'licenseDays': request.form.get('licenseDays'),
            'adjustment': float(request.form.get('adjustment'))
        }
        licencias.append(nueva_licencia)
        with open(json_path, 'w') as json_file:
            json.dump(data, json_file, indent=2)
        return redirect(url_for('licencias_especiales'))
    return render_template('add/addLicencias.html', 
                           error_message=None)

#TODO--------VALIDACION_DE_USUARIO_LOGUEADO--------------------------------------------------------
@app.before_request
def before_request():
    if request.endpoint != 'login' and 'usuario' not in session:
        return redirect(url_for('login'))

#TODO--------EXPORT_TO_EXCEL----------------------------------------------------------------------
@app.route('/export_excel', methods=['POST'])
def export_excel():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'nomina.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        nomina = data["Nomina"]
    
    excel_file = export_to_excel(nomina)
    return send_file(excel_file, 
                    as_attachment=True)
    
@app.route('/export_to_excel_azure', methods=['POST'])
def export_excel_azure():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(current_dir, 'data', 'json', 'nomina.json')
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)
        nomina = data["Nomina"]
    
    excel_file = export_to_excel_azure(nomina)
    return send_file(excel_file, 
                    as_attachment=True)

#TODO--------upload_nomina------------------------------------------------------------------------------
@app.route('/importar_nomina', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)        
        file = request.files['file']        
        if file.filename == '':
            return redirect(request.url)        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            return redirect(url_for('uploaded_file', filename=filename))    
    return render_template('upload.html')

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        return 'El archivo no existe.'
    json_data = convert_excel_to_json(file_path)
    json_filename = 'data/json/nomina.json'
    json_file_path = os.path.join(app.root_path, json_filename)
    with open(json_file_path, 'w') as json_file:
        json.dump(json_data, json_file, indent=2)
    return render_template('ok.html')

#TODO--------upload_licencias------------------------------------------------------------------------------
@app.route('/importar_licencia', methods=['GET', 'POST'])
def upload_licencia_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            return redirect(url_for('uploaded_licencia_file', 
                                    filename=filename))
    return render_template('upload_licencia.html')

@app.route('/uploads_licencia/<filename>')
def uploaded_licencia_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        return 'El archivo no existe.'
    json_data = convert_excel_licencia__to_json(file_path)
    json_filename = 'data/json/licencias.json'
    json_file_path = os.path.join(app.root_path, json_filename)
    with open(json_file_path, 'w') as json_file:
        json.dump(json_data, json_file, indent=2, default=serialize_datetime)
    return render_template('ok.html')


if __name__ == '__main__':    
    app.run()
