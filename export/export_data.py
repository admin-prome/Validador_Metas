from db.conection import DatabaseConnection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill

def export_to_excel(nomina):
    wb = Workbook()
    ws = wb.active
    bold_font = Font(bold=True)
    light_green = Color(rgb="C6EFCE")
    fill = PatternFill(start_color=light_green, end_color=light_green, fill_type="solid")
    columns = ["Legajo", "Nombre Ejecutivo", "Sucursal", "Categoria", "Meta Q", "Meta $", "Descripción Licencias", "Dias Licencencia", "Licencias Especiales", "Dias Licencias Especiales", "Es Tutor", "Tiene Progresion", "Ajuste Q, 1° Mes", "Ajuste Q, 2° Mes", "Ajuste Monto, 1° Mes", "Ajuste Monto, 2° Mes", "Ajuste Total Q", "Ajuste total Monto", "Observaciones"]
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font
        cell.fill = fill
    for row_data in nomina:
        ws.append(list(row_data))
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(len(str(cell.value)) for cell in ws[col_letter])
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width
    excel_file_path = 'data/export/Validacion_Metas.xlsx'
    excel_file = excel_file_path
    wb.save(excel_file)
    return excel_file

# Llamar a la función para exportar los datos a Excel
# export_to_excel()