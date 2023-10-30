from utils.metas import *
from utils.licencias import *
from utils.ajustes import *
from utils.otros_ajustes import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.utils import get_column_letter

def export_to_excel(nomina_data):
    wb = Workbook()
    ws = wb.active
    bold_font = Font(bold=True)
    light_green = Color(rgb="C6EFCE")
    fill = PatternFill(start_color=light_green, end_color=light_green, fill_type="solid")
    columns = ["Legajo", "Nombre Ejecutivo", "Sucursal", "Categoria", "Meta Q", "Meta $", "Descripción Licencias", "Dias Licencencia", "Licencias Especiales", "Dias Licencias Especiales", "Es Tutor", "Tiene Progresion", "Ajuste Q, 1° Mes", "Ajuste Q, 2° Mes", "Ajuste Monto, 1° Mes", "Ajuste Monto, 2° Mes", "Ajuste Total Q", "Ajuste total Monto", "Observaciones"]
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = bold_font
        cell.fill = fill
    for user in nomina_data:
        row_data = [
            user.get("employeeNumber", ""),
            user.get("fullName", ""),
            user.get("branch", ""),
            user.get("category", ""),
            get_meta_q(user.get("category", "")),
            get_meta_monto(user.get("category", "")),
            descripcion_licencias(user.get("employeeNumber", "")),
            cantidad_licencias(user.get("employeeNumber", "")),
            has_special_licence(user.get("employeeNumber", "")),
            has_special_licences_days(user.get("employeeNumber", "")),
            is_tutor(user.get("employeeNumber", "")),
            has_progresiones(user.get("employeeNumber", "")),
            ajuste_meta_q_mes_uno(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_q_mes_dos(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_monto_m1(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_meta_monto_m2(user.get("category", ""), user.get("employeeNumber", ""))/2,
            ajuste_total_q(user.get("category", ""), user.get("employeeNumber", "")),
            ajuste_total_monto(user.get("category", ""), user.get("employeeNumber", "")),
            user.get("Observations", "")
        ]
        ws.append(row_data)
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(len(str(cell.value)) for cell in ws[col_letter])
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width    
    # current_dir = os.path.dirname(os.path.abspath(__file__))
    # excel_file_path = os.path.join(current_dir, 'data', 'export', 'Validacion_Metas.xlsx')
    excel_file_path = os.path.join('data/export/Validacion_Metas.xlsx')
    excel_file = excel_file_path
    wb.save(excel_file)
    return excel_file